"""
note_visibility_report.py
─────────────────────────
Extracts note/subnote visibility settings from CaseWare Cloud SE author
templates and writes them to a formatted Excel spreadsheet.

SETUP
─────
1. Install dependencies:
       pip install requests openpyxl

2. Copy your full cookie string from the browser:
   - Open DevTools (F12) → Network tab
   - Click any "section/get" request → Headers tab
   - Scroll to Request Headers → find the "cookie:" line
   - Copy everything AFTER "cookie: " (the entire value on that line)

3. Set environment variable in PowerShell:
       $env:CW_COOKIES="<paste entire cookie string here>"

4. Edit the CONFIGURATION section below:
   - Set HOST and TENANT to match your environment
   - Add your templates to the TEMPLATES list (see comments)

5. Run:
       python note_visibility_report.py          # live data
       python note_visibility_report.py --mock   # test with sample data
       python note_visibility_report.py --debug  # print raw visibility JSON

HOW TO FIND engagementId AND documentId
───────────────────────────────────────
- engagementId: visible in the browser URL when viewing a template:
      .../e/eng/{engagementId}/...

- documentId: open DevTools → Network tab → filter by "section" →
  look for a POST to section/get → click it → Payload tab →
  copy the "value" field from the filter (e.g. "OfoGzA6HRqaeqdPVgSHg_w")
"""

import argparse
import os
import sys
import json
import logging
import re
from dataclasses import dataclass, astuple, fields
from html import unescape
from typing import Optional

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── CONFIGURATION ─────────────────────────────────────────────────────────────

HOST   = "https://ca.cwcloudpartner.com"
TENANT = "ca-develop"   # the path segment after the host in the URL

# Each entry: (display_name, engagement_id, document_id)
#   display_name  — label used in the Template Name column of the spreadsheet
#   engagement_id — from the browser URL: .../e/eng/{engagementId}/...
#   document_id   — from the section/get Payload in DevTools (the filter "value")
TEMPLATES = [
    ("Template A", "yK9KibLUQLOpAIEMo6XPXA", "OfoGzA6HRqaeqdPVgSHg_w"),
    # ("Template B", "<engagementId>", "<documentId>"),
]

OUTPUT_FILE = "note_visibility_report.xlsx"

# ── END CONFIGURATION ─────────────────────────────────────────────────────────

logging.basicConfig(
    level=os.environ.get("CW_LOG_LEVEL", "INFO").upper(),
    format="%(levelname)s  %(message)s",
)
log = logging.getLogger(__name__)

# Override → human-readable label for the Visibility Setting column
OVERRIDE_LABELS = {
    "default": "Use default settings",
    "show":    "Show",
    "hide":    "Hide",
}


@dataclass
class VisibilityRow:
    template_name:        str
    note_id:              str   # numeric position label
    note_title:           str   # e.g. "Basis of Accounting"
    subnote_id:           str   # blank if this IS the top-level note
    subnote_title:        str
    section_content:      str   # plain-text body of the section
    visibility_setting:   str   # Use default settings / Show / Hide
    visibility_behavior:  str   # "Hide when" / "Show when" / ""
    condition_group:      str   # e.g. "6-15 Financial statements optimiser"
    condition_name:       str   # e.g. "Statement of cash flows"
    expected_response:    str   # e.g. "Yes"


COLUMN_HEADERS = [f.name.replace("_", " ").title() for f in fields(VisibilityRow)]
COLUMN_WIDTHS   = [28, 10, 40, 12, 40, 60, 24, 20, 40, 45, 30]


# ── HTML STRIPPING ────────────────────────────────────────────────────────────

def strip_html(html: str) -> str:
    """Convert an HTML string to plain text by removing tags and unescaping entities."""
    if not html:
        return ""
    text = re.sub(r"<[^>]+>", " ", html)   # replace tags with a space
    text = unescape(text)                   # &amp; → &, &#160; → space, etc.
    text = re.sub(r"\s+", " ", text)        # collapse whitespace
    return text.strip()


# ── SESSION ───────────────────────────────────────────────────────────────────

def make_session() -> requests.Session:
    """Build a requests.Session pre-loaded with the browser cookie string."""
    cookies = os.environ.get("CW_COOKIES", "").strip()
    if not cookies:
        sys.exit(
            "ERROR: Missing environment variable CW_COOKIES.\n"
            "See the SETUP section at the top of this file."
        )

    session = requests.Session()
    session.headers.update({
        "Accept":       "application/json",
        "Content-Type": "application/json",
        "Cookie":       cookies,
    })
    return session


# ── DATA FETCHING ─────────────────────────────────────────────────────────────

def fetch_sections(session: requests.Session,
                   engagement_id: str,
                   document_id: str) -> list[dict]:
    """
    POST to section/get with a document-filter to retrieve every section
    belonging to the given document in one call.
    """
    url = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0/section/get"
    payload = {
        "filter": {
            "filter": {
                "node":  "=",
                "left":  {"node": "field", "kind": "section", "field": "document"},
                "right": {"node": "string", "value": document_id},
            }
        }
    }
    log.info("Fetching sections from %s", url)
    resp = session.post(url, json=payload, timeout=30)

    if resp.status_code == 401:
        sys.exit(
            "ERROR: 401 Unauthorised — your session cookies have expired.\n"
            "Re-copy CW_JSESSIONID and CW_SECID from the browser and try again."
        )
    if not resp.ok:
        sys.exit(
            f"ERROR: {resp.status_code} from {url}\n{resp.text[:500]}"
        )

    data = resp.json()
    log.debug("Raw response keys: %s", list(data.keys()) if isinstance(data, dict) else type(data).__name__)

    # The API may return a bare list, a {"objects": [...]} wrapper, or similar
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ("objects", "sections", "items", "results", "data"):
            if key in data and isinstance(data[key], list):
                return data[key]
        # Look for any list value (catches unexpected wrapper keys)
        for key, val in data.items():
            if isinstance(val, list):
                log.debug("Using response key '%s' (%d items)", key, len(val))
                return val
        # Single-object response wrapped as {"object": {...}} — wrap in a list
        if "object" in data and isinstance(data["object"], dict):
            log.debug("Single-object response — wrapping in list")
            return [data["object"]]

    log.warning("Unexpected response shape — returning empty list. Raw: %s", str(data)[:300])
    return []


def fetch_document_lookup(session: requests.Session, engagement_id: str) -> dict:
    """
    Fetch all documents in the engagement and return {id: "number name"} map.
    e.g. {"z2HFA1b5TIqLNYrLsb55ZA": "6-15 Financial statements optimiser"}
    """
    url = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0/document/get"
    try:
        resp = session.post(url, json={}, timeout=30)
        if not resp.ok:
            return {}
        data = resp.json()
        documents = data.get("objects", [])
        result = {}
        for doc in documents:
            did    = doc.get("id", "")
            number = doc.get("number", "")
            names  = doc.get("names") or {}
            name   = names.get("en", "") or doc.get("name", "")
            content = doc.get("content", "")
            if did:
                label = f"{number} {name}".strip() if number else name
                result[did] = label
                # checklistId in conditions points to the document's content ID, not its own ID
                if content:
                    result[content] = label
        log.info("  %d documents fetched for name lookup", len(result))
        return result
    except Exception as exc:
        log.warning("Could not fetch document list: %s", exc)
        return {}


def fetch_procedures_for_checklist(session: requests.Session,
                                   engagement_id: str,
                                   checklist_id: str) -> list[dict]:
    """Fetch all procedures belonging to a checklist."""
    url = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0/procedure/get"
    payload = {"filter": {"filter": {
        "node": "=",
        "left":  {"node": "field", "kind": "procedure", "field": "checklistId"},
        "right": {"node": "string", "value": checklist_id},
    }}}
    try:
        resp = session.post(url, json=payload, timeout=30)
        if not resp.ok:
            return []
        data = resp.json()
        return data.get("objects", []) if isinstance(data, dict) else []
    except Exception as exc:
        log.warning("Could not fetch procedures for checklist %s: %s", checklist_id, exc)
        return []


def fetch_procedure_by_id(session: requests.Session,
                          engagement_id: str,
                          procedure_id: str) -> Optional[dict]:
    """Fetch a single procedure by its ID."""
    url = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0/procedure/get"
    payload = {"filter": {"filter": {
        "node": "=",
        "left":  {"node": "field", "kind": "procedure", "field": "id"},
        "right": {"node": "string", "value": procedure_id},
    }}}
    try:
        resp = session.post(url, json=payload, timeout=15)
        if not resp.ok:
            return None
        data = resp.json()
        objects = data.get("objects", [])
        return objects[0] if objects else None
    except Exception as exc:
        log.debug("Could not fetch procedure %s: %s", procedure_id, exc)
        return None


def fetch_checklist_name(session: requests.Session,
                         engagement_id: str,
                         checklist_id: str) -> str:
    """
    Try to resolve a checklist ID to its human-readable name.
    Tries multiple strategies in order:
    0. Fetch the checklist as a procedure by its own ID (most reliable)
    1. Fetch the checklist object as a section by ID
    2. Walk up the procedure parentId chain to find the root title
    Returns empty string if not resolved.
    """
    # Strategy 0: the checklist IS a procedure — fetch it directly by its own ID
    proc = fetch_procedure_by_id(session, engagement_id, checklist_id)
    if proc:
        summary = proc.get("summaryNames") or {}
        name = summary.get("en", "") or next(iter(summary.values()), "")
        if not name:
            name = strip_html(proc.get("text", ""))
        if name:
            number = proc.get("number", "")
            return f"{number} {name}".strip() if number else name

    # Strategy 1a: fetch it as a document (document/get by id)
    for endpoint in ("document/get", "checklist/get", "workpaper/get"):
        url = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0/{endpoint}"
        for payload in (
            {"filter": {"filter": {"node": "=",
                "left": {"node": "field", "kind": endpoint.split("/")[0], "field": "id"},
                "right": {"node": "string", "value": checklist_id}}}},
            {"id": checklist_id},
        ):
            try:
                resp = session.post(url, json=payload, timeout=10)
                if not resp.ok:
                    continue
                data = resp.json()
                # Unwrap objects list or single object
                objects = data.get("objects") or []
                if not objects and "object" in data:
                    objects = [data["object"]]
                if not objects and "id" in data:
                    objects = [data]
                for obj in objects:
                    # Try 'name', 'title', 'titles', 'description' fields
                    for field in ("name", "names", "title", "titles", "description"):
                        val = obj.get(field)
                        if isinstance(val, dict):
                            val = val.get("en", "") or next(iter(val.values()), "")
                        if val and isinstance(val, str):
                            return val.strip()
            except Exception:
                continue

    # Strategy 1b: fetch it as a section by id field filter
    url = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0/section/get"
    payload = {"filter": {"filter": {
        "node": "=",
        "left":  {"node": "field", "kind": "section", "field": "id"},
        "right": {"node": "string", "value": checklist_id},
    }}}
    try:
        resp = session.post(url, json=payload, timeout=10)
        if resp.ok:
            data = resp.json()
            objects = data.get("objects", [])
            if objects:
                title = get_title(objects[0])
                if title:
                    return title
    except Exception:
        pass

    # Strategy 2: walk up the procedure parentId chain
    procs = fetch_procedures_for_checklist(session, engagement_id, checklist_id)
    if not procs:
        return ""

    seen = set()
    current = procs[0]
    while True:
        parent_id = current.get("parentId", "")
        if not parent_id or parent_id in seen:
            break
        seen.add(parent_id)
        parent = fetch_procedure_by_id(session, engagement_id, parent_id)
        if not parent:
            break
        if parent.get("checklistId", "") != checklist_id:
            break
        current = parent

    summary = current.get("summaryNames") or {}
    name = summary.get("en", "") or next(iter(summary.values()), "")
    if not name:
        name = strip_html(current.get("text", ""))
    return name


def build_id_lookup(session: requests.Session,
                    engagement_id: str,
                    sections: list[dict]) -> dict:
    """
    For every procedure ID referenced in visibility conditions:
      1. Fetch the procedure object
      2. Extract procedure text → {procedureId: "Statement of cash flows"}
      3. Extract response options from settings.responseSets[].responses[]
         → {responseId: "Yes"}, {responseId: "No"}, etc.
    """
    # Collect all unique procedure IDs referenced in conditions
    procedure_ids: set = set()
    for s in sections:
        for cond in (s.get("visibility") or {}).get("conditions") or []:
            _collect_procedure_ids_from_cond(cond, procedure_ids)

    log.info("  %d unique procedure ID(s) to resolve", len(procedure_ids))

    # Fetch all documents once — covers checklist name resolution
    doc_lookup = fetch_document_lookup(session, engagement_id)
    lookup: dict = dict(doc_lookup)  # seed with document names

    # Collect checklist IDs and resolve their names for the Condition Group column
    checklist_ids: set = set()
    for s in sections:
        for cond in (s.get("visibility") or {}).get("conditions") or []:
            _collect_checklist_ids_from_cond(cond, checklist_ids)

    # Build a quick index of sections by their 'document' field —
    # some sections have document = checklist_id when they belong to that checklist doc
    sections_by_doc: dict = {}
    for s in sections:
        doc_id = s.get("document", "")
        if doc_id:
            sections_by_doc.setdefault(doc_id, []).append(s)

    log.info("  Resolving %d checklist name(s) ...", len(checklist_ids))
    for cid in checklist_ids:
        # Skip if already resolved from fetch_document_lookup (e.g. "6-15 Financial statements optimiser")
        if cid in lookup:
            log.debug("  Checklist %s → '%s' (from document lookup)", cid[:12], lookup[cid])
            continue
        # Strategy 0: look for a section in already-fetched sections whose id = checklist_id
        # or whose document field = checklist_id and has a title
        name = ""
        # Check sections whose document matches this checklist
        for s in sections_by_doc.get(cid, []):
            t = get_title(s)
            if t:
                name = t
                break
        # Check if any fetched section has id = checklist_id
        if not name:
            for s in sections:
                if s.get("id") == cid and get_title(s):
                    name = get_title(s)
                    break
        if not name:
            name = fetch_checklist_name(session, engagement_id, cid)
        if name:
            lookup[cid] = name
            log.debug("  Checklist %s → '%s'", cid[:12], name)

    for idx, pid in enumerate(procedure_ids, start=1):
        proc = fetch_procedure_by_id(session, engagement_id, pid)
        if not proc:
            continue

        # Procedure name
        summary = proc.get("summaryNames") or {}
        proc_text = summary.get("en", "") or next(iter(summary.values()), "")
        if not proc_text:
            proc_text = strip_html(proc.get("text", ""))
        if pid and proc_text:
            lookup[pid] = proc_text

        # Response options live in settings.responseSets[].responses[]
        for rs in (proc.get("settings") or {}).get("responseSets") or []:
            for resp_opt in rs.get("responses") or []:
                rid  = resp_opt.get("id", "")
                name = resp_opt.get("name", "") or (resp_opt.get("names") or {}).get("en", "")
                if rid and name:
                    lookup[rid] = name

        if idx % 10 == 0:
            log.info("    ... %d / %d procedures resolved", idx, len(procedure_ids))

    log.info("  %d names resolved", len(lookup))
    return lookup


def _collect_procedure_ids_from_cond(cond: dict, ids: set) -> None:
    """Recursively collect procedureId values from a condition."""
    pid = (cond.get("procedureId") or {}).get("id")
    if pid:
        ids.add(pid)
    for sub in cond.get("conditions") or []:
        _collect_procedure_ids_from_cond(sub, ids)


def _collect_checklist_ids_from_cond(cond: dict, ids: set) -> None:
    """Recursively collect checklistId values from a condition."""
    cid = (cond.get("checklistId") or {}).get("id")
    if cid:
        ids.add(cid)
    for sub in cond.get("conditions") or []:
        _collect_checklist_ids_from_cond(sub, ids)


# ── TREE BUILDING ─────────────────────────────────────────────────────────────

def get_title(section: dict) -> str:
    """Return the display title of a section, trying multiple field names."""
    return (section.get("title")
            or section.get("titles", {}).get("en", "")
            or "").strip()


def find_nearest_titled_ancestor(section: dict, by_id: dict) -> Optional[dict]:
    """
    Walk up the parent chain and return the first ancestor that has a
    non-empty title.  Returns None if no titled ancestor exists in the set.
    """
    seen = {section.get("id", "")}
    parent_id = section.get("parent", "")
    while parent_id and parent_id in by_id:
        if parent_id in seen:
            break
        seen.add(parent_id)
        parent = by_id[parent_id]
        if get_title(parent):
            return parent
        parent_id = parent.get("parent", "")
    return None


def ordered_titled_sections(sections: list[dict]) -> list[dict]:
    """
    Return all sections that have a non-empty title, sorted so that
    parents always appear before their children (topological + order sort).
    """
    by_id    = {s["id"]: s for s in sections}
    id_set   = set(by_id)

    # Depth-first traversal from roots to preserve hierarchy order
    children: dict[str, list[dict]] = {}
    roots: list[dict] = []
    for s in sections:
        parent = s.get("parent", "")
        if parent in id_set:
            children.setdefault(parent, []).append(s)
        else:
            roots.append(s)

    def sort_key(s):
        return s.get("order", "")

    result: list[dict] = []

    def visit(s):
        if get_title(s):
            result.append(s)
        for child in sorted(children.get(s["id"], []), key=sort_key):
            visit(child)

    for root in sorted(roots, key=sort_key):
        visit(root)

    return result


# ── VISIBILITY PARSING ────────────────────────────────────────────────────────

def _override_label(vis: dict) -> str:
    raw = vis.get("override", "default")
    return OVERRIDE_LABELS.get(raw, raw)



def _resolve(lookup: dict, id_obj) -> str:
    """Resolve an id-object ({id: ..., authorId: ...}) to a human-readable name."""
    if not id_obj:
        return ""
    raw_id = id_obj.get("id", "") if isinstance(id_obj, dict) else str(id_obj)
    return lookup.get(raw_id, raw_id[:12])   # fall back to first 12 chars of ID


def _resolve_with_id(lookup: dict, id_obj) -> str:
    """Resolve an id-object to its human-readable name for display in column I."""
    return _resolve(lookup, id_obj)


def _flatten_conditions(conditions: list, lookup: dict,
                        group_label: str = "") -> list[tuple]:
    """
    Recursively flatten a conditions list into (group, condition, response) tuples.
    Handles both plain response conditions and nested condition_group objects.
    """
    rows = []
    group_counter = [0]   # mutable counter shared across recursion

    for cond in conditions:
        ctype = cond.get("type", "")

        if ctype == "response":
            checklist = _resolve_with_id(lookup, cond.get("checklistId"))
            procedure = _resolve(lookup, cond.get("procedureId"))
            response  = _resolve(lookup, cond.get("responseId"))
            rows.append((group_label or checklist, procedure, response))

        elif ctype == "condition_group":
            group_counter[0] += 1
            nested = cond.get("conditions") or []
            # Derive a group label from the checklist of the first nested condition
            first_checklist = ""
            if nested:
                first_checklist = _resolve_with_id(lookup, nested[0].get("checklistId"))
            label = f"Condition Group {group_counter[0]}: {first_checklist}".strip(": ")
            rows.extend(_flatten_conditions(nested, lookup, group_label=label))

        else:
            # Unknown condition type — show raw type so nothing is silently dropped
            rows.append((group_label, ctype, json.dumps(cond)[:80]))

    return rows


def parse_visibility(section: dict,
                     lookup: dict,
                     debug: bool = False) -> list[dict]:
    """
    Return a list of dicts representing condition rows for this section.
    Each dict contains visibility_setting, visibility_behavior,
    condition_group, condition_name, expected_response.
    One row per condition; one empty row if there are no conditions.
    """
    vis = section.get("visibility") or {}

    if debug and vis:
        log.debug("RAW visibility for '%s':\n%s",
                  section.get("title", section.get("id")),
                  json.dumps(vis, indent=2))

    override   = _override_label(vis)
    conditions = vis.get("conditions") or []

    # Determine hide/show behavior label
    if override == "default" or not conditions:
        behavior = ""
    elif override in ("hide", "always_hide"):
        behavior = "Hide when"
    else:
        behavior = "Show when"

    base = dict(
        visibility_setting = override,
        visibility_behavior = behavior,
    )

    if not conditions:
        return [{**base, "condition_group": "", "condition_name": "", "expected_response": ""}]

    rows = []
    for group, name, resp in _flatten_conditions(conditions, lookup):
        rows.append({**base,
                     "condition_group":   group,
                     "condition_name":    name,
                     "expected_response": resp})
    return rows if rows else [{**base,
                                "condition_group": "", "condition_name": "",
                                "expected_response": ""}]


def section_rows(template_name: str,
                 section: dict,
                 by_id: dict,
                 note_counter: dict,
                 lookup: dict,
                 debug: bool) -> list[VisibilityRow]:
    """
    Build VisibilityRow objects for a single titled section.
    Finds the nearest titled ancestor to use as Note Title;
    the section itself becomes Subnote Title.
    If no titled ancestor exists, the section is the note (Subnote Title blank).
    """
    ancestor  = find_nearest_titled_ancestor(section, by_id)
    sec_title = get_title(section)

    if ancestor:
        note_title    = get_title(ancestor)
        subnote_title = sec_title
        note_key      = ancestor["id"]
        sub_key       = section["id"]
    else:
        note_title    = sec_title
        subnote_title = ""
        note_key      = section["id"]
        sub_key       = ""

    # Assign stable numeric IDs for alternating row colours
    if note_key not in note_counter:
        note_counter[note_key] = str(len(note_counter) + 1)
    note_id    = note_counter[note_key]
    subnote_id = sub_key[:8] if sub_key else ""

    raw_html = (section.get("specification") or {}).get("content", "")
    content  = strip_html(raw_html)

    vis_rows = parse_visibility(section, lookup=lookup, debug=debug)
    result   = []
    for vr in vis_rows:
        result.append(VisibilityRow(
            template_name       = template_name,
            note_id             = note_id,
            note_title          = note_title,
            subnote_id          = subnote_id,
            subnote_title       = subnote_title,
            section_content     = content,
            visibility_setting  = vr["visibility_setting"],
            visibility_behavior = vr["visibility_behavior"],
            condition_group     = vr["condition_group"],
            condition_name      = vr["condition_name"],
            expected_response   = vr["expected_response"],
        ))
    return result


# ── EXCEL WRITING ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="FF1F3864")  # dark navy
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
FILL_A      = PatternFill("solid", fgColor="FFFFFFFF")  # white
FILL_B      = PatternFill("solid", fgColor="FFD9E1F2")  # light blue
ALIGN_TOP   = Alignment(vertical="top", wrap_text=False)


def write_excel(rows: list[VisibilityRow], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Note Visibility"

    # Header row
    ws.append(COLUMN_HEADERS)
    for col_idx, (cell, width) in enumerate(
            zip(ws[1], COLUMN_WIDTHS), start=1):
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = ALIGN_TOP
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes           = "A2"
    ws.auto_filter.ref        = f"A1:{get_column_letter(len(COLUMN_HEADERS))}1"

    # Data rows with alternating fill per (template, note_id, subnote_id)
    fill_toggle = False
    prev_key    = None

    for row in rows:
        key = (row.template_name, row.note_id, row.subnote_id)
        if key != prev_key:
            fill_toggle = not fill_toggle
            prev_key    = key

        ws.append(list(astuple(row)))
        fill = FILL_B if fill_toggle else FILL_A
        for col_idx, cell in enumerate(ws[ws.max_row], start=1):
            cell.fill      = fill
            # Wrap text in the Section Content column (column 6) only
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 6))

    try:
        wb.save(output_path)
    except PermissionError:
        sys.exit(
            f"ERROR: Cannot write to '{output_path}'.\n"
            "If the file is open in Excel, close it and run again."
        )


# ── MOCK DATA ─────────────────────────────────────────────────────────────────

def load_mock_sections(document_id: str) -> list[dict]:
    """
    Hardcoded sections that exercise every parser path:
      - Top-level notes with no subnotes
      - Subnotes with no conditions (Use default settings)
      - Subnotes with conditions (Show)
      - Subnotes overridden to Hide
    """
    return [
        # Note a
        {"id": "note_a", "parent": document_id, "order": "1",
         "title": "Basis of Accounting", "kind": "section",
         "visibility": {"override": "default", "normallyVisible": True,
                        "allConditionsNeeded": False, "conditions": []}},
        # Subnote a-i  (no conditions)
        {"id": "sub_ai", "parent": "note_a", "order": "1a",
         "title": "Basis of Accounting - no cash flow", "kind": "section",
         "visibility": {"override": "default", "normallyVisible": True,
                        "allConditionsNeeded": False, "conditions": []}},
        # Subnote a-ii  (Show, 2 conditions OR-logic)
        {"id": "sub_aii", "parent": "note_a", "order": "1b",
         "title": "Basis of Accounting - cash flow", "kind": "section",
         "visibility": {
             "override": "show", "normallyVisible": True,
             "allConditionsNeeded": False,
             "conditions": [
                 {"type": "response", "checklistId": "checklist_1",
                  "procedureId": "proc_1", "expectedResponse": "Yes",
                  "_label": "Statement of cash flows"},
                 {"type": "response", "checklistId": "checklist_1",
                  "procedureId": "proc_2", "expectedResponse": "Yes",
                  "_label": "Would you like to use a condensed format"},
             ]}},
        # Note b
        {"id": "note_b", "parent": document_id, "order": "2",
         "title": "Revenue Recognition", "kind": "section",
         "visibility": {"override": "default", "normallyVisible": True,
                        "allConditionsNeeded": False, "conditions": []}},
        # Subnote b-i  (Hide)
        {"id": "sub_bi", "parent": "note_b", "order": "2a",
         "title": "Revenue - accrual basis", "kind": "section",
         "visibility": {"override": "hide", "normallyVisible": False,
                        "allConditionsNeeded": True,
                        "conditions": [
                            {"type": "consolidation", "consolidated": True,
                             "_label": "Consolidated entity"}
                        ]}},
    ]


# ── ENDPOINT PROBE (temporary diagnostic tool) ───────────────────────────────

def probe_endpoints(session: requests.Session, engagement_id: str) -> None:
    """
    Try a set of candidate endpoints with the known checklist ID to find
    where procedure/response names are stored. Prints a summary and exits.
    """
    known_checklist_id = "z2HFA1b5TIqLNYrLsb55ZA"
    known_procedure_id = "HZKIXcUORaOovqz2d567RA"
    known_response_id  = "OpsVFmLFSHajjVr-hN5Duw"
    base = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0"

    def mk_filter(kind, field, value):
        return {"filter": {"filter": {"node": "=",
                "left": {"node": "field", "kind": kind, "field": field},
                "right": {"node": "string", "value": value}}}}

    # Probe document/get for the known checklist ID
    print("\n=== DOCUMENT/GET for checklist ID ===")
    for path in ("document/get",):
        url = f"{base}/{path}"
        for payload in (
            mk_filter("document", "id", known_checklist_id),
            {"id": known_checklist_id},
            {},
        ):
            resp = session.post(url, json=payload, timeout=10)
            try:
                data = resp.json()
            except Exception:
                print(f"  {resp.status_code} {path} payload={list(payload.keys())} — not JSON")
                continue
            count = data.get("count", "?") if isinstance(data, dict) else "?"
            print(f"  {resp.status_code} {path} payload={list(payload.keys())} count={count}")
            objects = data.get("objects", []) if isinstance(data, dict) else []
            if objects:
                obj = objects[0]
                print(f"    keys: {list(obj.keys())[:12]}")
                for f in ("name", "names", "title", "titles", "description", "label"):
                    v = obj.get(f)
                    if v:
                        print(f"    {f}: {str(v)[:120]}")
                break
    print("=====================================\n")

    # Try fetching the FSO checklist as a document via section/get
    print("=== CHECKLIST AS DOCUMENT (section/get) ===")
    url = f"{base}/section/get"
    payload = {"filter": {"filter": {"node": "=",
        "left": {"node": "field", "kind": "section", "field": "document"},
        "right": {"node": "string", "value": known_checklist_id}}}}
    resp = session.post(url, json=payload, timeout=10)
    data = resp.json() if resp.ok else {}
    objects = data.get("objects", [])
    print(f"  status={resp.status_code}  count={data.get('count', '?')}  objects={len(objects)}")
    # Look for any section whose id matches our known_response_id
    matches = [o for o in objects if o.get("id") == known_response_id]
    print(f"  responseId match found: {bool(matches)}")
    if objects:
        # Print a sample to see the structure
        sample = objects[0]
        print(f"  first obj keys: {list(sample.keys())[:10]}")
        print(f"  first obj title: {sample.get('title','')}")
    # Also look for any section with our procedure ID
    proc_matches = [o for o in objects if o.get("id") == known_procedure_id]
    print(f"  procedureId match found: {bool(proc_matches)}")
    print("===========================================\n")

    candidates = [
        # Checklist by id field
        ("checklist/get", mk_filter("checklist", "id", known_checklist_id)),
        # Checklist - fetch all
        ("checklist/get", {}),
        # Procedure by id
        ("procedure/get", {"id": known_procedure_id}),
        # Procedure filtered by checklist field (various field names)
        ("procedure/get", mk_filter("procedure", "checklist", known_checklist_id)),
        ("procedure/get", mk_filter("procedure", "checklistId", known_checklist_id)),
        ("procedure/get", mk_filter("procedure", "document", known_checklist_id)),
        # Procedure - fetch all (no filter)
        ("procedure/get", {}),
        # Response by id
        ("response/get",  {"id": known_response_id}),
        # Response filtered
        ("response/get",  mk_filter("response", "procedure", known_procedure_id)),
        ("response/get",  {}),
        # Other possible endpoints
        ("checklistitem/get", {}),
        ("question/get",  {}),
        ("answer/get",    {}),
        ("responseset/get", {}),
        ("responseset/get", {"id": known_response_id}),
        ("responseset/get", mk_filter("responseset", "id", known_response_id)),
        ("responseSet/get", {}),
        ("responseOption/get", {}),
    ]

    for path, payload in candidates:
        url = f"{base}/{path}"
        try:
            resp = session.post(url, json=payload, timeout=10)
            try:
                data = resp.json()
            except Exception:
                print(f"  {resp.status_code}  {path:35s}  (not JSON)")
                continue
            count  = data.get("count", "?") if isinstance(data, dict) else len(data) if isinstance(data, list) else "?"
            errors = data.get("errors", []) if isinstance(data, dict) else []
            keys   = list(data.keys())[:6] if isinstance(data, dict) else "list"
            print(f"  {resp.status_code}  {path:35s}  count={count}  keys={keys}  errors={errors[:1]}")
            if resp.ok and count not in (0, "?", "0"):
                objects = data.get("objects", [data] if "id" in data else [])
                if objects:
                    obj = objects[0]
                    print(f"         first obj keys: {list(obj.keys())[:12]}")
                    title = obj.get('title') or obj.get('name') or obj.get('text') or obj.get('content','')
                    print(f"         title/name: {str(title)[:100]}")
        except Exception as e:
            print(f"  ERR  {path:35s}  {e}")

    sys.exit(0)


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export CaseWare note visibility settings to Excel."
    )
    parser.add_argument("--mock",   action="store_true",
                        help="Use built-in sample data (no API calls)")
    parser.add_argument("--debug",  action="store_true",
                        help="Print raw visibility JSON for each section")
    parser.add_argument("--probe",  action="store_true",
                        help="Test candidate endpoints to find procedure/response names")
    parser.add_argument("--output", default=OUTPUT_FILE,
                        help=f"Output file path (default: {OUTPUT_FILE})")
    args = parser.parse_args()

    if args.debug:
        log.setLevel(logging.DEBUG)

    if not TEMPLATES:
        sys.exit("ERROR: No templates configured. Edit the TEMPLATES list.")

    session = None if args.mock else make_session()

    if args.probe:
        _, engagement_id, _ = TEMPLATES[0]
        print(f"\nProbing endpoints for engagement: {engagement_id}\n")
        probe_endpoints(session, engagement_id)

    all_rows: list[VisibilityRow] = []

    for (template_name, engagement_id, document_id) in TEMPLATES:
        log.info("Processing template: %s", template_name)

        if args.mock:
            sections = load_mock_sections(document_id)
        else:
            sections = fetch_sections(session, engagement_id, document_id)

        if not sections:
            log.warning("No sections returned for template '%s' — skipping.",
                        template_name)
            continue

        log.info("  %d sections fetched", len(sections))
        by_id        = {s["id"]: s for s in sections}
        titled       = ordered_titled_sections(sections)
        note_counter: dict = {}
        log.info("  %d titled sections found", len(titled))

        # Build lookup: fetch each condition ID individually to resolve to names
        lookup: dict = {}
        if not args.mock:
            lookup = build_id_lookup(session, engagement_id, sections)
            log.info("  %d names resolved", len(lookup))

        for sec in titled:
            all_rows.extend(section_rows(template_name, sec, by_id, note_counter, lookup, args.debug))

    if not all_rows:
        sys.exit("No data collected. Check your template configuration and cookies.")

    write_excel(all_rows, args.output)
    log.info("Wrote %d rows to %s", len(all_rows), args.output)
    print(f"\nDone — {len(all_rows)} rows written to: {args.output}")


if __name__ == "__main__":
    main()
