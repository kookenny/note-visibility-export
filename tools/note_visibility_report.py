"""
note_visibility_report.py
─────────────────────────
Extracts note/subnote visibility settings from CaseWare Cloud SE author
templates and writes them to a formatted Excel spreadsheet.

SETUP
─────
1. Install dependencies:
       pip install requests openpyxl

2. Copy your full cookie string from the browser:
   - Open DevTools (F12) → Network tab
   - Click any "section/get" request → Headers tab
   - Scroll to Request Headers → find the "cookie:" line
   - Copy everything AFTER "cookie: " (the entire value on that line)

3. Set environment variable in PowerShell:
       $env:CW_COOKIES="<paste entire cookie string here>"

4. Edit the CONFIGURATION section below:
   - Set HOST and TENANT to match your environment
   - Add your templates to the TEMPLATES list (see comments)

5. Run:
       python note_visibility_report.py          # live data
       python note_visibility_report.py --mock   # test with sample data
       python note_visibility_report.py --debug  # print raw visibility JSON

HOW TO FIND engagementId AND documentId
───────────────────────────────────────
- engagementId: visible in the browser URL when viewing a template:
      .../e/eng/{engagementId}/...

- documentId: open DevTools → Network tab → filter by "section" →
  look for a POST to section/get → click it → Payload tab →
  copy the "value" field from the filter (e.g. "OfoGzA6HRqaeqdPVgSHg_w")
"""

import argparse
import io
import os
import sys
import json
import logging
import re
from dataclasses import dataclass, astuple, fields
from html import unescape
from typing import Optional

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── CONFIGURATION ─────────────────────────────────────────────────────────────

HOST   = "https://ca.cwcloudpartner.com"
TENANT = "ca-develop"   # the path segment after the host in the URL

# Each entry: (display_name, engagement_id, document_id)
#   display_name  — label used in the Template Name column of the spreadsheet
#   engagement_id — from the browser URL: .../e/eng/{engagementId}/...
#   document_id   — from the section/get Payload in DevTools (the filter "value")
TEMPLATES = [
    ("Template A", "yK9KibLUQLOpAIEMo6XPXA", "OfoGzA6HRqaeqdPVgSHg_w"),
    # ("Template B", "<engagementId>", "<documentId>"),
]

OUTPUT_FILE = ".tmp/note_visibility_report.xlsx"

# ── END CONFIGURATION ─────────────────────────────────────────────────────────

logging.basicConfig(
    level=os.environ.get("CW_LOG_LEVEL", "INFO").upper(),
    format="%(levelname)s  %(message)s",
)
log = logging.getLogger(__name__)

# Override → human-readable label for the Visibility Setting column
OVERRIDE_LABELS = {
    "default": "Use default settings",
    "show":    "Show",
    "hide":    "Hide",
}


@dataclass
class VisibilityRow:
    note_number:          str   # e.g. "1", "2" — top-level note number
    subnote_letter:       str   # e.g. "a", "b" — subnote position letter
    item_numeral:         str   # e.g. "i", "ii" — third-level roman numeral
    note_group_title:     str   # e.g. "Summary of accounting policies"
    note_title:           str   # e.g. "Basis of Accounting"
    subnote_title:        str   # e.g. "Basis of Accounting - no cash flow"
    content_title:        str   # e.g. "Cash - no cash flow" (leaf section title)
    section_content:      str   # plain-text body of the section
    visibility:           str   # "Hide when all" / "Show when any" / "Hide" / "Show" / "Use default settings"
    condition_group:      str   # e.g. "6-15 Financial statements optimiser (any)" (first row of group only)
    condition_name:       str   # e.g. "Statement of cash flows"
    expected_response:    str   # e.g. "Yes"
    components:           str   # e.g. "NFP, Charities and foundation"


COLUMN_HEADERS = [f.name.replace("_", " ").title() for f in fields(VisibilityRow)]
COLUMN_WIDTHS   = [10, 10, 10, 40, 40, 40, 40, 60, 28, 40, 45, 30, 35]


# ── HTML STRIPPING ────────────────────────────────────────────────────────────

def strip_html(html: str, formula_map: dict[str, str] | None = None) -> str:
    """Convert an HTML string to plain text by removing tags and unescaping entities.
    Placeholder spans are wrapped in (( )).
    Formula/dynamic-text spans are resolved via *formula_map* and wrapped in [[ ]]."""
    if not html:
        return ""
    # Wrap placeholder span contents in (( )) before stripping tags
    text = re.sub(
        r'<span[^>]*\bplaceholder="[^"]*"[^>]*>(.*?)</span>',
        r"((\1))",
        html,
    )
    # Resolve dynamic-text formula spans → [[calculated value]]
    if formula_map:
        def _resolve_formula(m: re.Match) -> str:
            fid = m.group(1)
            val = formula_map.get(fid, "")
            return f"[[{val}]]" if val else ""
        text = re.sub(
            r'<span[^>]*\bformula="([^"]*)"[^>]*>.*?</span>',
            _resolve_formula,
            text,
        )
    text = re.sub(r"<[^>]+>", " ", text)   # replace tags with a space
    text = unescape(text)                   # &amp; → &, &#160; → space, etc.
    text = re.sub(r"\s+", " ", text)        # collapse whitespace
    return text.strip()


def build_formula_map(section: dict) -> dict[str, str]:
    """Build a {referenceId: calculated_value} map from a section's attachables."""
    result = {}
    for a in (section.get("attachables") or {}).values():
        ref_id = a.get("referenceId")
        calc = (a.get("calculated") or "").strip()
        if ref_id and calc:
            result[ref_id] = calc
    return result


# ── SESSION ───────────────────────────────────────────────────────────────────

def _env_prefix_from_host(host: str) -> str:
    """Derive an env-var prefix from the hostname: 'ca.cwcloudpartner.com' → 'CA'."""
    hostname = host.replace("https://", "").replace("http://", "").split("/")[0]
    return hostname.split(".")[0].upper()


def _obtain_bearer_token(env_prefix: str = "") -> str | None:
    """Exchange CW_CLIENT_ID + CW_CLIENT_SECRET for a Bearer token via OAuth.

    Looks for prefixed env vars first (e.g. CW_CA_CLIENT_ID) then falls back
    to the generic CW_CLIENT_ID / CW_CLIENT_SECRET.
    """
    client_id, client_secret = "", ""
    if env_prefix:
        client_id     = os.environ.get(f"CW_{env_prefix}_CLIENT_ID", "").strip()
        client_secret = os.environ.get(f"CW_{env_prefix}_CLIENT_SECRET", "").strip()
    if not client_id or not client_secret:
        client_id     = os.environ.get("CW_CLIENT_ID", "").strip()
        client_secret = os.environ.get("CW_CLIENT_SECRET", "").strip()
    if not client_id or not client_secret:
        return None
    url = f"{HOST}/{TENANT}/ms/caseware-cloud/api/v1/auth/token"
    resp = requests.post(url, json={
        "ClientId": client_id, "ClientSecret": client_secret, "Language": "en",
    }, headers={"Accept": "application/json", "Content-Type": "application/json"},
       timeout=15)
    resp.raise_for_status()
    token = resp.json().get("Token")
    if token:
        log.info("Authenticated via OAuth (Bearer token)")
    return token


def make_session(env_prefix: str = "") -> requests.Session:
    """Build a requests.Session using OAuth (preferred) or browser cookies."""
    session = requests.Session()
    session.headers.update({
        "Accept":       "application/json",
        "Content-Type": "application/json",
    })

    # Try OAuth first
    token = _obtain_bearer_token(env_prefix)
    if token:
        session.headers["Authorization"] = f"Bearer {token}"
        return session

    # Fall back to cookie auth
    cookies = os.environ.get("CW_COOKIES", "").strip()
    if not cookies:
        raise RuntimeError(
            "No auth credentials found. "
            "Set CW_CLIENT_ID + CW_CLIENT_SECRET for OAuth, "
            "or CW_COOKIES for cookie auth."
        )
    session.headers["Cookie"] = cookies
    log.info("Authenticated via browser cookies")
    return session


# ── DATA FETCHING ─────────────────────────────────────────────────────────────

def fetch_sections(session: requests.Session,
                   engagement_id: str,
                   document_id: str) -> list[dict]:
    """
    POST to section/get with a document-filter to retrieve every section
    belonging to the given document in one call.
    """
    url = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0/section/get"
    payload = {
        "filter": {
            "filter": {
                "node":  "=",
                "left":  {"node": "field", "kind": "section", "field": "document"},
                "right": {"node": "string", "value": document_id},
            }
        }
    }
    log.info("Fetching sections from %s", url)
    resp = session.post(url, json=payload, timeout=30)

    if resp.status_code == 401:
        raise RuntimeError(
            "401 Unauthorised — your session cookies have expired. "
            "Re-copy CW_JSESSIONID and CW_SECID from the browser and try again."
        )
    if not resp.ok:
        raise RuntimeError(
            f"{resp.status_code} from {url}\n{resp.text[:500]}"
        )

    data = resp.json()
    log.debug("Raw response keys: %s", list(data.keys()) if isinstance(data, dict) else type(data).__name__)

    # The API may return a bare list, a {"objects": [...]} wrapper, or similar
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ("objects", "sections", "items", "results", "data"):
            if key in data and isinstance(data[key], list):
                return data[key]
        # Look for any list value (catches unexpected wrapper keys)
        for key, val in data.items():
            if isinstance(val, list):
                log.debug("Using response key '%s' (%d items)", key, len(val))
                return val
        # Single-object response wrapped as {"object": {...}} — wrap in a list
        if "object" in data and isinstance(data["object"], dict):
            log.debug("Single-object response — wrapping in list")
            return [data["object"]]

    log.warning("Unexpected response shape — returning empty list. Raw: %s", str(data)[:300])
    return []


def fetch_component_lookup(session: requests.Session,
                           engagement_id: str) -> dict[str, str]:
    """Fetch all tags with subKind='component' and return {tag_id: name} map."""
    url = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0/tag/get"
    log.info("Fetching component tags from %s", url)
    resp = session.post(url, json={}, timeout=30)
    if not resp.ok:
        log.warning("tag/get returned %s — skipping components", resp.status_code)
        return {}
    tags = resp.json().get("objects", [])
    return {t["id"]: t.get("name", "") for t in tags if t.get("subKind") == "component"}


def build_section_components(sections: list[dict],
                             component_lookup: dict[str, str]) -> dict[str, str]:
    """Build {section_id: 'CompA, CompB'} from section.tagging.component."""
    result = {}
    for s in sections:
        comp_map = (s.get("tagging") or {}).get("component")
        if not comp_map:
            continue
        names = []
        for tag_ids in comp_map.values():
            for tid in tag_ids:
                name = component_lookup.get(tid, "")
                if name:
                    names.append(name)
        if names:
            result[s["id"]] = ", ".join(sorted(set(names)))
    return result


def fetch_document_lookup(session: requests.Session, engagement_id: str) -> dict:
    """
    Fetch all documents in the engagement and return {id: "number name"} map.
    e.g. {"z2HFA1b5TIqLNYrLsb55ZA": "6-15 Financial statements optimiser"}
    """
    url = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0/document/get"
    try:
        resp = session.post(url, json={}, timeout=30)
        if not resp.ok:
            return {}
        data = resp.json()
        documents = data.get("objects", [])
        result = {}
        for doc in documents:
            did    = doc.get("id", "")
            number = doc.get("number", "")
            names  = doc.get("names") or {}
            name   = names.get("en", "") or doc.get("name", "")
            content = doc.get("content", "")
            if did:
                label = f"{number} {name}".strip() if number else name
                result[did] = label
                # checklistId in conditions points to the document's content ID, not its own ID
                if content:
                    result[content] = label
        log.info("  %d documents fetched for name lookup", len(result))
        return result
    except Exception as exc:
        log.warning("Could not fetch document list: %s", exc)
        return {}


def fetch_procedures_for_checklist(session: requests.Session,
                                   engagement_id: str,
                                   checklist_id: str) -> list[dict]:
    """Fetch all procedures belonging to a checklist."""
    url = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0/procedure/get"
    payload = {"filter": {"filter": {
        "node": "=",
        "left":  {"node": "field", "kind": "procedure", "field": "checklistId"},
        "right": {"node": "string", "value": checklist_id},
    }}}
    try:
        resp = session.post(url, json=payload, timeout=30)
        if not resp.ok:
            return []
        data = resp.json()
        return data.get("objects", []) if isinstance(data, dict) else []
    except Exception as exc:
        log.warning("Could not fetch procedures for checklist %s: %s", checklist_id, exc)
        return []


def fetch_procedure_by_id(session: requests.Session,
                          engagement_id: str,
                          procedure_id: str) -> Optional[dict]:
    """Fetch a single procedure by its ID."""
    url = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0/procedure/get"
    payload = {"filter": {"filter": {
        "node": "=",
        "left":  {"node": "field", "kind": "procedure", "field": "id"},
        "right": {"node": "string", "value": procedure_id},
    }}}
    try:
        resp = session.post(url, json=payload, timeout=15)
        if not resp.ok:
            return None
        data = resp.json()
        objects = data.get("objects", [])
        return objects[0] if objects else None
    except Exception as exc:
        log.debug("Could not fetch procedure %s: %s", procedure_id, exc)
        return None


def fetch_checklist_name(session: requests.Session,
                         engagement_id: str,
                         checklist_id: str) -> str:
    """
    Try to resolve a checklist ID to its human-readable name.
    Tries multiple strategies in order:
    0. Fetch the checklist as a procedure by its own ID (most reliable)
    1. Fetch the checklist object as a section by ID
    2. Walk up the procedure parentId chain to find the root title
    Returns empty string if not resolved.
    """
    # Strategy 0: the checklist IS a procedure — fetch it directly by its own ID
    proc = fetch_procedure_by_id(session, engagement_id, checklist_id)
    if proc:
        summary = proc.get("summaryNames") or {}
        name = summary.get("en", "") or next(iter(summary.values()), "")
        if not name:
            name = strip_html(proc.get("text", ""))
        if name:
            number = proc.get("number", "")
            return f"{number} {name}".strip() if number else name

    # Strategy 1a: fetch it as a document (document/get by id)
    for endpoint in ("document/get", "checklist/get", "workpaper/get"):
        url = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0/{endpoint}"
        for payload in (
            {"filter": {"filter": {"node": "=",
                "left": {"node": "field", "kind": endpoint.split("/")[0], "field": "id"},
                "right": {"node": "string", "value": checklist_id}}}},
            {"id": checklist_id},
        ):
            try:
                resp = session.post(url, json=payload, timeout=10)
                if not resp.ok:
                    continue
                data = resp.json()
                # Unwrap objects list or single object
                objects = data.get("objects") or []
                if not objects and "object" in data:
                    objects = [data["object"]]
                if not objects and "id" in data:
                    objects = [data]
                for obj in objects:
                    # Try 'name', 'title', 'titles', 'description' fields
                    for field in ("name", "names", "title", "titles", "description"):
                        val = obj.get(field)
                        if isinstance(val, dict):
                            val = val.get("en", "") or next(iter(val.values()), "")
                        if val and isinstance(val, str):
                            return val.strip()
            except Exception:
                continue

    # Strategy 1b: fetch it as a section by id field filter
    url = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0/section/get"
    payload = {"filter": {"filter": {
        "node": "=",
        "left":  {"node": "field", "kind": "section", "field": "id"},
        "right": {"node": "string", "value": checklist_id},
    }}}
    try:
        resp = session.post(url, json=payload, timeout=10)
        if resp.ok:
            data = resp.json()
            objects = data.get("objects", [])
            if objects:
                title = get_title(objects[0])
                if title:
                    return title
    except Exception:
        pass

    # Strategy 2: walk up the procedure parentId chain
    procs = fetch_procedures_for_checklist(session, engagement_id, checklist_id)
    if not procs:
        return ""

    seen = set()
    current = procs[0]
    while True:
        parent_id = current.get("parentId", "")
        if not parent_id or parent_id in seen:
            break
        seen.add(parent_id)
        parent = fetch_procedure_by_id(session, engagement_id, parent_id)
        if not parent:
            break
        if parent.get("checklistId", "") != checklist_id:
            break
        current = parent

    summary = current.get("summaryNames") or {}
    name = summary.get("en", "") or next(iter(summary.values()), "")
    if not name:
        name = strip_html(current.get("text", ""))
    return name


def build_id_lookup(session: requests.Session,
                    engagement_id: str,
                    sections: list[dict]) -> dict:
    """
    For every procedure ID referenced in visibility conditions:
      1. Fetch the procedure object
      2. Extract procedure text → {procedureId: "Statement of cash flows"}
      3. Extract response options from settings.responseSets[].responses[]
         → {responseId: "Yes"}, {responseId: "No"}, etc.
    """
    # Collect all unique procedure IDs referenced in conditions
    procedure_ids: set = set()
    for s in sections:
        for cond in (s.get("visibility") or {}).get("conditions") or []:
            _collect_procedure_ids_from_cond(cond, procedure_ids)

    log.info("  %d unique procedure ID(s) to resolve", len(procedure_ids))

    # Fetch all documents once — covers checklist name resolution
    doc_lookup = fetch_document_lookup(session, engagement_id)
    lookup: dict = dict(doc_lookup)  # seed with document names

    # Collect checklist IDs and resolve their names for the Condition Group column
    checklist_ids: set = set()
    for s in sections:
        for cond in (s.get("visibility") or {}).get("conditions") or []:
            _collect_checklist_ids_from_cond(cond, checklist_ids)

    # Build a quick index of sections by their 'document' field —
    # some sections have document = checklist_id when they belong to that checklist doc
    sections_by_doc: dict = {}
    for s in sections:
        doc_id = s.get("document", "")
        if doc_id:
            sections_by_doc.setdefault(doc_id, []).append(s)

    log.info("  Resolving %d checklist name(s) ...", len(checklist_ids))
    for cid in checklist_ids:
        # Skip if already resolved from fetch_document_lookup (e.g. "6-15 Financial statements optimiser")
        if cid in lookup:
            log.debug("  Checklist %s → '%s' (from document lookup)", cid[:12], lookup[cid])
            continue
        # Strategy 0: look for a section in already-fetched sections whose id = checklist_id
        # or whose document field = checklist_id and has a title
        name = ""
        # Check sections whose document matches this checklist
        for s in sections_by_doc.get(cid, []):
            t = get_title(s)
            if t:
                name = t
                break
        # Check if any fetched section has id = checklist_id
        if not name:
            for s in sections:
                if s.get("id") == cid and get_title(s):
                    name = get_title(s)
                    break
        if not name:
            name = fetch_checklist_name(session, engagement_id, cid)
        if name:
            lookup[cid] = name
            log.debug("  Checklist %s → '%s'", cid[:12], name)

    for idx, pid in enumerate(procedure_ids, start=1):
        proc = fetch_procedure_by_id(session, engagement_id, pid)
        if not proc:
            continue

        # Procedure name
        summary = proc.get("summaryNames") or {}
        proc_text = summary.get("en", "") or next(iter(summary.values()), "")
        if not proc_text:
            proc_text = strip_html(proc.get("text", ""))
        if pid and proc_text:
            lookup[pid] = proc_text

        # Response options live in settings.responseSets[].responses[]
        for rs in (proc.get("settings") or {}).get("responseSets") or []:
            for resp_opt in rs.get("responses") or []:
                rid  = resp_opt.get("id", "")
                name = resp_opt.get("name", "") or (resp_opt.get("names") or {}).get("en", "")
                if rid and name:
                    lookup[rid] = name

        if idx % 10 == 0:
            log.info("    ... %d / %d procedures resolved", idx, len(procedure_ids))

    log.info("  %d names resolved", len(lookup))
    return lookup


def _collect_procedure_ids_from_cond(cond: dict, ids: set) -> None:
    """Recursively collect procedureId values from a condition."""
    pid = (cond.get("procedureId") or {}).get("id")
    if pid:
        ids.add(pid)
    for sub in cond.get("conditions") or []:
        _collect_procedure_ids_from_cond(sub, ids)


def _collect_checklist_ids_from_cond(cond: dict, ids: set) -> None:
    """Recursively collect checklistId values from a condition."""
    cid = (cond.get("checklistId") or {}).get("id")
    if cid:
        ids.add(cid)
    for sub in cond.get("conditions") or []:
        _collect_checklist_ids_from_cond(sub, ids)


# ── TREE BUILDING ─────────────────────────────────────────────────────────────

# Section types that are structural containers, not user-visible content levels
_CONTAINER_TYPES = {"heading", "note", "settings", "toc"}

def get_ancestor_chain(section: dict, by_id: dict) -> list:
    """
    Return the full list of titled ancestor titles from the root down to
    (and including) this section, excluding structural container sections
    (type in _CONTAINER_TYPES), e.g.:
      ["Summary of accounting policies", "Basis of accounting",
       "Basis of Accounting - no cash flow"]
    """
    chain = []
    current = section
    visited: set = set()
    while current:
        sid = current.get("id", "")
        if sid in visited:
            break
        visited.add(sid)
        sec_type = current.get("type", "")
        title = (current.get("title") or current.get("titles", {}).get("en", "") or "").strip()
        if title and sec_type not in _CONTAINER_TYPES:
            chain.append(title)
        parent_id = current.get("parent", "")
        current = by_id.get(parent_id)
    chain.reverse()
    return chain


def get_title(section: dict) -> str:
    """
    Return the display title of a section, stripped of HTML tags.
    note-type sections store their real title in specification.title;
    the top-level title field is always the generic string "Note".
    """
    raw = (section.get("title")
           or section.get("titles", {}).get("en", "")
           or "").strip()
    title = strip_html(raw) if raw else ""
    # For note sections the meaningful title is in specification.title
    if not title or title == "Note":
        spec = section.get("specification") or {}
        spec_title = (spec.get("title")
                      or (spec.get("titles") or {}).get("en", "")
                      or "")
        if spec_title:
            title = strip_html(spec_title)
    return title


def get_note_hierarchy(section: dict, by_id: dict) -> tuple[str, str, str]:
    """
    Walk up the parent chain (including note types) and return
    (note_group, note, subnote) — the 3 nearest titled non-structural ancestors.
    Structural types (heading, settings, toc) are skipped; note types are included.
    Returns empty strings for levels that don't exist.
    """
    _STRUCTURAL = {"heading", "settings", "toc"}
    ancestors: list[str] = []
    parent_id = section.get("parent", "")
    visited = {section.get("id", "")}
    while parent_id and parent_id in by_id:
        if parent_id in visited:
            break
        visited.add(parent_id)
        parent = by_id[parent_id]
        title = get_title(parent)
        if title and parent.get("type", "") not in _STRUCTURAL:
            ancestors.append(title)
        parent_id = parent.get("parent", "")
    # ancestors[0] = immediate titled parent, [1] = grandparent, [2] = great-grandparent
    subnote = ancestors[0] if len(ancestors) >= 1 else ""
    note    = ancestors[1] if len(ancestors) >= 2 else ""
    group   = ancestors[2] if len(ancestors) >= 3 else ""
    return group, note, subnote


def find_nearest_titled_ancestor(section: dict, by_id: dict) -> Optional[dict]:
    """
    Walk up the parent chain and return the first ancestor that has a
    non-empty title AND is not a structural container type.
    Returns None if no such ancestor exists.
    """
    seen = {section.get("id", "")}
    parent_id = section.get("parent", "")
    while parent_id and parent_id in by_id:
        if parent_id in seen:
            break
        seen.add(parent_id)
        parent = by_id[parent_id]
        if get_title(parent) and parent.get("type", "") not in _CONTAINER_TYPES:
            return parent
        parent_id = parent.get("parent", "")
    return None


def ordered_titled_sections(sections: list[dict]) -> list[dict]:
    """
    Return all content sections (non-container types) that have a non-empty
    title, sorted so that parents always appear before their children.
    Container types (note, heading, settings, toc) are excluded — their
    visibility is duplicated on the child content sections.
    """
    by_id    = {s["id"]: s for s in sections}
    id_set   = set(by_id)

    # Depth-first traversal from roots to preserve hierarchy order
    children: dict[str, list[dict]] = {}
    roots: list[dict] = []
    for s in sections:
        parent = s.get("parent", "")
        if parent in id_set:
            children.setdefault(parent, []).append(s)
        else:
            roots.append(s)

    def sort_key(s):
        return s.get("order", "")

    result: list[dict] = []

    def visit(s):
        if get_title(s) and s.get("type", "") not in _CONTAINER_TYPES:
            result.append(s)
        for child in sorted(children.get(s["id"], []), key=sort_key):
            visit(child)

    for root in sorted(roots, key=sort_key):
        visit(root)

    return result


def _to_roman(n: int) -> str:
    """Convert a positive integer to a lowercase roman numeral string."""
    vals = [(1000,'m'),(900,'cm'),(500,'d'),(400,'cd'),(100,'c'),(90,'xc'),
            (50,'l'),(40,'xl'),(10,'x'),(9,'ix'),(5,'v'),(4,'iv'),(1,'i')]
    result = []
    for val, numeral in vals:
        while n >= val:
            result.append(numeral)
            n -= val
    return "".join(result)


def compute_note_numbers(sections: list[dict]) -> dict[str, tuple[str, str, str]]:
    """
    Compute note numbers (1, 2, 3...), subnote letters (a, b, c...),
    and item numerals (i, ii, iii...) based on position among siblings,
    matching the CaseWare UI numbering.

    Structure: heading → note (top-level, numbered) → note (subnote, lettered)
               → note (item, roman numeral) → content (leaf).

    Returns {section_id: (note_number, subnote_letter, item_numeral)}.
    """
    by_id = {s["id"]: s for s in sections}

    # Group children by parent, sorted by order
    children_by_parent: dict[str, list[dict]] = {}
    for s in sections:
        parent = s.get("parent", "")
        children_by_parent.setdefault(parent, []).append(s)
    for pid in children_by_parent:
        children_by_parent[pid].sort(key=lambda s: s.get("order", ""))

    # Level 1 — top-level note containers: parent is a heading
    top_note_num: dict[str, int] = {}     # note_id → 1-based number
    heading_ids = {s["id"] for s in sections if s.get("type") == "heading"}
    for hid in heading_ids:
        idx = 0
        for s in children_by_parent.get(hid, []):
            if s.get("type") == "note":
                idx += 1
                top_note_num[s["id"]] = idx

    # Level 2 — subnote containers: parent is a top-level note
    sub_note_letter: dict[str, str] = {}  # note_id → "a", "b", ...
    for top_id in top_note_num:
        idx = 0
        for s in children_by_parent.get(top_id, []):
            if s.get("type") == "note":
                letter = chr(ord('a') + idx) if idx < 26 else str(idx + 1)
                sub_note_letter[s["id"]] = letter
                idx += 1

    # Level 3 — item containers: parent is a subnote
    item_numeral: dict[str, str] = {}     # note_id → "i", "ii", ...
    for sub_id in sub_note_letter:
        idx = 0
        for s in children_by_parent.get(sub_id, []):
            if s.get("type") == "note":
                idx += 1
                item_numeral[s["id"]] = _to_roman(idx)

    # For each content section, walk up the note chain to find all three levels
    result: dict[str, tuple[str, str, str]] = {}
    for s in sections:
        if s.get("type", "") in _CONTAINER_TYPES or not get_title(s):
            continue
        sid = s["id"]
        current = s.get("parent", "")
        visited = {sid}
        found_item = ""
        found_sub = ""
        found_top = ""
        while current and current in by_id:
            if current in visited:
                break
            visited.add(current)
            if current in item_numeral and not found_item:
                found_item = item_numeral[current]
            if current in sub_note_letter and not found_sub:
                found_sub = sub_note_letter[current]
            if current in top_note_num:
                found_top = str(top_note_num[current])
                break
            current = by_id[current].get("parent", "")

        result[sid] = (found_top, found_sub, found_item)

    return result


# ── VISIBILITY PARSING ────────────────────────────────────────────────────────

def _effective_visibility(section: dict, by_id: dict) -> dict:
    """
    Return the visibility dict to use for a section.
    If the section has its own conditions, return its visibility.
    Otherwise walk up the parent chain to find the nearest ancestor
    (typically a [note] container) that carries conditions.
    """
    vis = section.get("visibility") or {}
    if vis.get("conditions"):
        return vis
    parent_id = section.get("parent", "")
    visited = {section.get("id", "")}
    while parent_id and parent_id in by_id:
        if parent_id in visited:
            break
        visited.add(parent_id)
        parent = by_id[parent_id]
        pvis = parent.get("visibility") or {}
        if pvis.get("conditions"):
            return pvis
        parent_id = parent.get("parent", "")
    return vis


def _override_label(vis: dict) -> str:
    raw = vis.get("override", "default")
    return OVERRIDE_LABELS.get(raw, raw)



def _resolve(lookup: dict, id_obj) -> str:
    """Resolve an id-object ({id: ..., authorId: ...}) to a human-readable name."""
    if not id_obj:
        return ""
    raw_id = id_obj.get("id", "") if isinstance(id_obj, dict) else str(id_obj)
    return lookup.get(raw_id, raw_id[:12])   # fall back to first 12 chars of ID


def _resolve_with_id(lookup: dict, id_obj) -> str:
    """Resolve an id-object to its human-readable name for display in column I."""
    return _resolve(lookup, id_obj)


def _flatten_conditions(conditions: list, lookup: dict,
                        group_label: str = "",
                        is_group: bool = False) -> list[tuple]:
    """
    Recursively flatten a conditions list into (group, condition, response, is_group) tuples.
    Handles both plain response conditions and nested condition_group objects.
    """
    rows = []

    for cond in conditions:
        ctype = cond.get("type", "")

        if ctype == "response":
            checklist = _resolve_with_id(lookup, cond.get("checklistId"))
            procedure = _resolve(lookup, cond.get("procedureId"))
            response  = _resolve(lookup, cond.get("responseId"))
            rows.append((group_label or checklist, procedure, response, is_group))

        elif ctype == "condition_group":
            nested = cond.get("conditions") or []
            group_all = cond.get("allConditionsNeeded", False)
            qualifier = "all" if group_all else "any"
            # Derive a group label from the checklist of the first nested condition
            first_checklist = ""
            if nested:
                first_checklist = _resolve_with_id(lookup, nested[0].get("checklistId"))
            label = f"{first_checklist} ({qualifier})" if first_checklist else f"({qualifier})"
            rows.extend(_flatten_conditions(nested, lookup, group_label=label, is_group=True))

        else:
            # Unknown condition type — show raw type so nothing is silently dropped
            rows.append((group_label, ctype, json.dumps(cond)[:80], is_group))

    return rows


def parse_visibility(vis: dict,
                     lookup: dict,
                     debug: bool = False,
                     debug_label: str = "") -> list[dict]:
    """
    Return a list of dicts representing condition rows for a visibility dict.
    Each dict contains visibility, condition_group, condition_name, expected_response.
    One row per condition; one empty row if there are no conditions.

    The visibility dict may come from the section itself or from an ancestor
    (via _effective_visibility).
    """
    if debug and vis:
        log.debug("RAW visibility for '%s':\n%s",
                  debug_label, json.dumps(vis, indent=2))

    raw_override     = vis.get("override", "default")
    conditions       = vis.get("conditions") or []
    normally_visible = vis.get("normallyVisible", True)
    all_needed       = vis.get("allConditionsNeeded", False)
    quantifier       = "all" if all_needed else "any"

    # Derive the single merged visibility label.
    # "show"/"hide" overrides force visibility regardless of conditions.
    # "default" + conditions: direction comes from normallyVisible:
    #   normallyVisible=false → "Show when" (normally hidden; show when conditions met)
    #   normallyVisible=true  → "Hide when" (normally visible; hide when conditions met)
    # quantifier from allConditionsNeeded: "any" or "all"
    if raw_override == "show" and not conditions:
        visibility = "Show"
    elif raw_override == "hide" and not conditions:
        visibility = "Hide"
    elif raw_override in ("show", "hide") and conditions:
        direction = "Hide" if raw_override == "hide" else "Show"
        visibility = f"{direction} when {quantifier}"
    elif conditions:
        direction = "Show" if not normally_visible else "Hide"
        visibility = f"{direction} when {quantifier}"
    else:
        visibility = "Use default settings"

    if not conditions:
        return [{"visibility": visibility, "condition_group": "", "condition_name": "", "expected_response": ""}]

    rows = []
    for group, name, resp, is_group in _flatten_conditions(conditions, lookup):
        rows.append({"visibility":       visibility,
                     "condition_group":  group,
                     "condition_name":   name,
                     "expected_response": resp,
                     "is_group":         is_group})
    return rows if rows else [{"visibility": visibility, "condition_group": "", "condition_name": "", "expected_response": ""}]


def section_rows(section: dict,
                 by_id: dict,
                 children_by_parent: dict,
                 lookup: dict,
                 debug: bool,
                 section_components: dict[str, str] | None = None,
                 note_numbers: dict[str, str] | None = None) -> list[VisibilityRow]:
    """
    Build VisibilityRow objects for a single titled section.
    Walks the full ancestor chain (including note types) to populate
    Note Group, Note, and Subnote hierarchy columns.
    Content from untitled child sections (Text/body) is merged into Section Content.
    """
    group_title, note_title, subnote_title = get_note_hierarchy(section, by_id)
    content_title = get_title(section)
    note_number, subnote_letter, item_numeral = (note_numbers or {}).get(section["id"], ("", "", ""))

    raw_html = (section.get("specification") or {}).get("content", "")
    fmap     = build_formula_map(section)
    content  = strip_html(raw_html, fmap)

    # Merge content from untitled child sections (Heading → Text pattern)
    untitled_children = sorted(
        [c for c in children_by_parent.get(section["id"], []) if not get_title(c)],
        key=lambda s: s.get("order", ""),
    )
    for child in untitled_children:
        child_fmap = build_formula_map(child)
        child_text = strip_html((child.get("specification") or {}).get("content", ""), child_fmap)
        if child_text:
            content = f"{content}\n{child_text}".strip() if content else child_text

    vis = _effective_visibility(section, by_id)
    debug_label = section.get("title", section.get("id", ""))
    vis_rows = parse_visibility(vis, lookup=lookup, debug=debug, debug_label=debug_label)

    comp_label = (section_components or {}).get(section["id"], "")

    # Show visibility only on first row; show condition_group only on first row of each group
    prev_group = None
    result     = []
    for i, vr in enumerate(vis_rows):
        raw_group   = vr["condition_group"]
        is_group    = vr.get("is_group", False)
        prefix      = "[Group] " if is_group else ""
        group_label = (f"{prefix}{raw_group}" if raw_group else "") if raw_group != prev_group else ""
        prev_group  = vr["condition_group"]
        first = (i == 0)
        result.append(VisibilityRow(
            note_number      = note_number      if first else "",
            subnote_letter   = subnote_letter   if first else "",
            item_numeral     = item_numeral     if first else "",
            note_group_title = group_title      if first else "",
            note_title       = note_title       if first else "",
            subnote_title    = subnote_title    if first else "",
            content_title    = content_title    if first else "",
            section_content  = (content or "{Review dynamic table/analysis graph in product}") if first else "",
            visibility       = vr["visibility"] if first else "",
            condition_group  = group_label,
            condition_name   = vr["condition_name"],
            expected_response = vr["expected_response"],
            components       = comp_label       if first else "",
        ))
    return result


# ── EXCEL WRITING ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="FF1F3864")  # dark navy
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
FILL_A      = PatternFill("solid", fgColor="FFFFFFFF")  # white
FILL_B      = PatternFill("solid", fgColor="FFD9E1F2")  # light blue
ALIGN_TOP   = Alignment(vertical="top", wrap_text=False)


def write_excel(rows: list[VisibilityRow], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Note Visibility"

    # Header row
    ws.append(COLUMN_HEADERS)
    for col_idx, (cell, width) in enumerate(
            zip(ws[1], COLUMN_WIDTHS), start=1):
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = ALIGN_TOP
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes           = "A2"
    ws.auto_filter.ref        = f"A1:{get_column_letter(len(COLUMN_HEADERS))}1"

    # Data rows with alternating fill per content title (column D)
    fill_toggle = False
    prev_key    = None

    for row in rows:
        # Non-first condition rows have blank identifiers; only update key when populated
        key = (row.note_number, row.subnote_letter, row.item_numeral, row.note_group_title, row.note_title, row.subnote_title, row.content_title)
        if key != ("", "", "", "", "", "", "") and key != prev_key:
            fill_toggle = not fill_toggle
            prev_key    = key

        ws.append(list(astuple(row)))
        fill = FILL_B if fill_toggle else FILL_A
        for col_idx, cell in enumerate(ws[ws.max_row], start=1):
            cell.fill      = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    try:
        wb.save(output_path)
    except PermissionError:
        raise RuntimeError(
            f"Cannot write to '{output_path}'. "
            "If the file is open in Excel, close it and run again."
        )


def write_excel_to_bytes(rows: list[VisibilityRow]) -> bytes:
    """Write the Excel workbook to an in-memory buffer and return raw bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Note Visibility"

    ws.append(COLUMN_HEADERS)
    for col_idx, (cell, width) in enumerate(
            zip(ws[1], COLUMN_WIDTHS), start=1):
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = ALIGN_TOP
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMN_HEADERS))}1"

    fill_toggle = False
    prev_key    = None

    for row in rows:
        key = (row.note_number, row.subnote_letter, row.item_numeral,
               row.note_group_title, row.note_title, row.subnote_title,
               row.content_title)
        if key != ("", "", "", "", "", "", "") and key != prev_key:
            fill_toggle = not fill_toggle
            prev_key    = key

        ws.append(list(astuple(row)))
        fill = FILL_B if fill_toggle else FILL_A
        for col_idx, cell in enumerate(ws[ws.max_row], start=1):
            cell.fill      = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_report_bytes(
    engagement_id: str,
    document_id: str,
    template_name: str = "Report",
    host: str | None = None,
    tenant: str | None = None,
) -> bytes:
    """
    Generate an Excel report for one template and return it as bytes.
    Called by the web frontend.  Raises on error instead of sys.exit().
    """
    global HOST, TENANT
    old_host, old_tenant = HOST, TENANT
    if host:
        HOST = host
    if tenant:
        TENANT = tenant
    try:
        env_prefix = _env_prefix_from_host(host) if host else ""
        session = make_session(env_prefix)
        sections = fetch_sections(session, engagement_id, document_id)
        if not sections:
            raise ValueError(
                f"No sections returned for engagement={engagement_id}, "
                f"document={document_id}"
            )

        by_id   = {s["id"]: s for s in sections}
        titled  = ordered_titled_sections(sections)

        children_by_parent: dict = {}
        for s in sections:
            pid = s.get("parent", "")
            if pid in by_id:
                children_by_parent.setdefault(pid, []).append(s)

        lookup = build_id_lookup(session, engagement_id, sections)
        component_lookup = fetch_component_lookup(session, engagement_id)
        section_components = build_section_components(sections, component_lookup)
        note_numbers = compute_note_numbers(sections)

        all_rows: list[VisibilityRow] = []
        for sec in titled:
            all_rows.extend(section_rows(
                sec, by_id, children_by_parent, lookup,
                False, section_components, note_numbers,
            ))

        if not all_rows:
            raise ValueError("No data collected. Check engagement/document IDs.")

        return write_excel_to_bytes(all_rows)
    finally:
        HOST, TENANT = old_host, old_tenant


# ── MOCK DATA ─────────────────────────────────────────────────────────────────

def load_mock_sections(document_id: str) -> list[dict]:
    """
    Hardcoded sections that exercise every parser path:
      - Top-level notes with no subnotes
      - Subnotes with no conditions (Use default settings)
      - Subnotes with conditions (Show)
      - Subnotes overridden to Hide
    """
    return [
        # Note container a (type=note — structural wrapper)
        {"id": "note_ctr_a", "parent": document_id, "order": "1",
         "title": "Note", "type": "note",
         "visibility": {"override": "default", "normallyVisible": True,
                        "allConditionsNeeded": False, "conditions": []}},
        # Note a — content child of note container
        {"id": "note_a", "parent": "note_ctr_a", "order": "1",
         "title": "Basis of Accounting", "type": "content",
         "visibility": {"override": "default", "normallyVisible": True,
                        "allConditionsNeeded": False, "conditions": []}},
        # Subnote a-i  (no conditions)
        {"id": "sub_ai", "parent": "note_a", "order": "1a",
         "title": "Basis of Accounting - no cash flow", "type": "content",
         "visibility": {"override": "default", "normallyVisible": True,
                        "allConditionsNeeded": False, "conditions": []}},
        # Subnote a-ii  (Show, 2 conditions OR-logic)
        {"id": "sub_aii", "parent": "note_a", "order": "1b",
         "title": "Basis of Accounting - cash flow", "type": "content",
         "visibility": {
             "override": "show", "normallyVisible": True,
             "allConditionsNeeded": False,
             "conditions": [
                 {"type": "response", "checklistId": "checklist_1",
                  "procedureId": "proc_1", "expectedResponse": "Yes",
                  "_label": "Statement of cash flows"},
                 {"type": "response", "checklistId": "checklist_1",
                  "procedureId": "proc_2", "expectedResponse": "Yes",
                  "_label": "Would you like to use a condensed format"},
             ]}},
        # Note container b
        {"id": "note_ctr_b", "parent": document_id, "order": "2",
         "title": "Note", "type": "note",
         "visibility": {"override": "default", "normallyVisible": True,
                        "allConditionsNeeded": False, "conditions": []}},
        # Note b
        {"id": "note_b", "parent": "note_ctr_b", "order": "1",
         "title": "Revenue Recognition", "type": "content",
         "visibility": {"override": "default", "normallyVisible": True,
                        "allConditionsNeeded": False, "conditions": []}},
        # Subnote b-i  (Hide)
        {"id": "sub_bi", "parent": "note_b", "order": "2a",
         "title": "Revenue - accrual basis", "type": "content",
         "visibility": {"override": "hide", "normallyVisible": False,
                        "allConditionsNeeded": True,
                        "conditions": [
                            {"type": "consolidation", "consolidated": True,
                             "_label": "Consolidated entity"}
                        ]}},
    ]


# ── ENDPOINT PROBE (temporary diagnostic tool) ───────────────────────────────

def probe_endpoints(session: requests.Session, engagement_id: str) -> None:
    """
    Try a set of candidate endpoints with the known checklist ID to find
    where procedure/response names are stored. Prints a summary and exits.
    """
    known_checklist_id = "z2HFA1b5TIqLNYrLsb55ZA"
    known_procedure_id = "HZKIXcUORaOovqz2d567RA"
    known_response_id  = "OpsVFmLFSHajjVr-hN5Duw"
    base = f"{HOST}/{TENANT}/e/eng/{engagement_id}/api/v1.12.0"

    def mk_filter(kind, field, value):
        return {"filter": {"filter": {"node": "=",
                "left": {"node": "field", "kind": kind, "field": field},
                "right": {"node": "string", "value": value}}}}

    # Probe document/get for the known checklist ID
    print("\n=== DOCUMENT/GET for checklist ID ===")
    for path in ("document/get",):
        url = f"{base}/{path}"
        for payload in (
            mk_filter("document", "id", known_checklist_id),
            {"id": known_checklist_id},
            {},
        ):
            resp = session.post(url, json=payload, timeout=10)
            try:
                data = resp.json()
            except Exception:
                print(f"  {resp.status_code} {path} payload={list(payload.keys())} — not JSON")
                continue
            count = data.get("count", "?") if isinstance(data, dict) else "?"
            print(f"  {resp.status_code} {path} payload={list(payload.keys())} count={count}")
            objects = data.get("objects", []) if isinstance(data, dict) else []
            if objects:
                obj = objects[0]
                print(f"    keys: {list(obj.keys())[:12]}")
                for f in ("name", "names", "title", "titles", "description", "label"):
                    v = obj.get(f)
                    if v:
                        print(f"    {f}: {str(v)[:120]}")
                break
    print("=====================================\n")

    # Try fetching the FSO checklist as a document via section/get
    print("=== CHECKLIST AS DOCUMENT (section/get) ===")
    url = f"{base}/section/get"
    payload = {"filter": {"filter": {"node": "=",
        "left": {"node": "field", "kind": "section", "field": "document"},
        "right": {"node": "string", "value": known_checklist_id}}}}
    resp = session.post(url, json=payload, timeout=10)
    data = resp.json() if resp.ok else {}
    objects = data.get("objects", [])
    print(f"  status={resp.status_code}  count={data.get('count', '?')}  objects={len(objects)}")
    # Look for any section whose id matches our known_response_id
    matches = [o for o in objects if o.get("id") == known_response_id]
    print(f"  responseId match found: {bool(matches)}")
    if objects:
        # Print a sample to see the structure
        sample = objects[0]
        print(f"  first obj keys: {list(sample.keys())[:10]}")
        print(f"  first obj title: {sample.get('title','')}")
    # Also look for any section with our procedure ID
    proc_matches = [o for o in objects if o.get("id") == known_procedure_id]
    print(f"  procedureId match found: {bool(proc_matches)}")
    print("===========================================\n")

    candidates = [
        # Checklist by id field
        ("checklist/get", mk_filter("checklist", "id", known_checklist_id)),
        # Checklist - fetch all
        ("checklist/get", {}),
        # Procedure by id
        ("procedure/get", {"id": known_procedure_id}),
        # Procedure filtered by checklist field (various field names)
        ("procedure/get", mk_filter("procedure", "checklist", known_checklist_id)),
        ("procedure/get", mk_filter("procedure", "checklistId", known_checklist_id)),
        ("procedure/get", mk_filter("procedure", "document", known_checklist_id)),
        # Procedure - fetch all (no filter)
        ("procedure/get", {}),
        # Response by id
        ("response/get",  {"id": known_response_id}),
        # Response filtered
        ("response/get",  mk_filter("response", "procedure", known_procedure_id)),
        ("response/get",  {}),
        # Other possible endpoints
        ("checklistitem/get", {}),
        ("question/get",  {}),
        ("answer/get",    {}),
        ("responseset/get", {}),
        ("responseset/get", {"id": known_response_id}),
        ("responseset/get", mk_filter("responseset", "id", known_response_id)),
        ("responseSet/get", {}),
        ("responseOption/get", {}),
    ]

    for path, payload in candidates:
        url = f"{base}/{path}"
        try:
            resp = session.post(url, json=payload, timeout=10)
            try:
                data = resp.json()
            except Exception:
                print(f"  {resp.status_code}  {path:35s}  (not JSON)")
                continue
            count  = data.get("count", "?") if isinstance(data, dict) else len(data) if isinstance(data, list) else "?"
            errors = data.get("errors", []) if isinstance(data, dict) else []
            keys   = list(data.keys())[:6] if isinstance(data, dict) else "list"
            print(f"  {resp.status_code}  {path:35s}  count={count}  keys={keys}  errors={errors[:1]}")
            if resp.ok and count not in (0, "?", "0"):
                objects = data.get("objects", [data] if "id" in data else [])
                if objects:
                    obj = objects[0]
                    print(f"         first obj keys: {list(obj.keys())[:12]}")
                    title = obj.get('title') or obj.get('name') or obj.get('text') or obj.get('content','')
                    print(f"         title/name: {str(title)[:100]}")
        except Exception as e:
            print(f"  ERR  {path:35s}  {e}")

    sys.exit(0)


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export CaseWare note visibility settings to Excel."
    )
    parser.add_argument("--mock",   action="store_true",
                        help="Use built-in sample data (no API calls)")
    parser.add_argument("--debug",  action="store_true",
                        help="Print raw visibility JSON for each section")
    parser.add_argument("--probe",  action="store_true",
                        help="Test candidate endpoints to find procedure/response names")
    parser.add_argument("--output", default=OUTPUT_FILE,
                        help=f"Output file path (default: {OUTPUT_FILE})")
    parser.add_argument("--dump-sections", action="store_true",
                        help="Save raw sections JSON to .tmp/sections_dump.json and exit")
    args = parser.parse_args()

    if args.debug:
        log.setLevel(logging.DEBUG)

    if not TEMPLATES:
        sys.exit("ERROR: No templates configured. Edit the TEMPLATES list.")

    try:
        session = None if args.mock else make_session()
    except RuntimeError as e:
        sys.exit(f"ERROR: {e}")

    if args.probe:
        _, engagement_id, _ = TEMPLATES[0]
        print(f"\nProbing endpoints for engagement: {engagement_id}\n")
        probe_endpoints(session, engagement_id)

    if args.dump_sections:
        _, engagement_id, document_id = TEMPLATES[0]
        sections = fetch_sections(session, engagement_id, document_id)
        dump_path = ".tmp/sections_dump.json"
        with open(dump_path, "w", encoding="utf-8") as f:
            json.dump(sections, f, indent=2)
        print(f"Dumped {len(sections)} sections to {dump_path}")
        sys.exit(0)

    all_rows: list[VisibilityRow] = []

    for (template_name, engagement_id, document_id) in TEMPLATES:
        log.info("Processing template: %s", template_name)

        if args.mock:
            sections = load_mock_sections(document_id)
        else:
            sections = fetch_sections(session, engagement_id, document_id)

        if not sections:
            log.warning("No sections returned for template '%s' — skipping.",
                        template_name)
            continue

        log.info("  %d sections fetched", len(sections))
        by_id        = {s["id"]: s for s in sections}
        titled       = ordered_titled_sections(sections)

        log.info("  %d titled sections found", len(titled))

        # Map parent id → all child sections (including untitled Text/body sections)
        children_by_parent: dict = {}
        for s in sections:
            pid = s.get("parent", "")
            if pid in by_id:
                children_by_parent.setdefault(pid, []).append(s)

        # Build lookup: fetch each condition ID individually to resolve to names
        lookup: dict = {}
        section_components: dict[str, str] = {}
        if not args.mock:
            lookup = build_id_lookup(session, engagement_id, sections)
            log.info("  %d names resolved", len(lookup))
            component_lookup = fetch_component_lookup(session, engagement_id)
            section_components = build_section_components(sections, component_lookup)
            log.info("  %d sections have components", len(section_components))

        note_numbers = compute_note_numbers(sections)

        for sec in titled:
            all_rows.extend(section_rows(sec, by_id, children_by_parent, lookup, args.debug, section_components, note_numbers))

    if not all_rows:
        sys.exit("No data collected. Check your template configuration and cookies.")

    write_excel(all_rows, args.output)
    log.info("Wrote %d rows to %s", len(all_rows), args.output)
    print(f"\nDone — {len(all_rows)} rows written to: {args.output}")


if __name__ == "__main__":
    main()
